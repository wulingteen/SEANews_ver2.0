"""
Excel 生成服務
從新聞文件內容中解析新聞列表並生成 Excel 報告
"""
import os
import re
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openai import OpenAI


def extract_country_from_content(content: str, fallback_name: str = "") -> str:
    """
    從文章內容中判斷國家（使用規則式 URL 網域匹配，不使用 LLM）
    
    Args:
        content: 文章內容
        fallback_name: 備用的文件名稱
        
    Returns:
        國家名稱（中文）
    """
    # 區域英文轉中文映射
    region_to_chinese = {
        'Vietnam': '越南',
        'Thailand': '泰國',
        'Singapore': '新加坡',
        'Philippines': '菲律賓',
        'Cambodia': '柬埔寨',
        'Indonesia': '印尼',
        'Malaysia': '馬來西亞',
        'Myanmar': '緬甸',
        'Laos': '寮國',
        'Southeast Asia': '東南亞',
    }
    
    # 網域到國家的映射（從 TRUSTED_NEWS_SOURCES 提取）
    domain_to_country = {
        'viet-jo.com': '越南',
        'cafef.vn': '越南',
        'vnexpress.net': '越南',
        'vietnamfinance.vn': '越南',
        'vir.com.vn': '越南',
        'vietnambiz.vn': '越南',
        'tapchikinhtetaichinh.vn': '越南',
        'bangkokpost.com': '泰國',
        'techsauce.co': '泰國',
        'fintechnews.sg': '新加坡',
        'fintechnews.ph': '菲律賓',
        'khmertimeskh.com': '柬埔寨',
        'cc-times.com': '柬埔寨',
        'phnompenhpost.com': '柬埔寨',
        'dealstreetasia.com': '東南亞',
        'techinasia.com': '東南亞',
        'asia.nikkei.com': '東南亞',
        'heaptalk.com': '東南亞',
    }
    
    # 嘗試從內容中的 URL 提取國家
    import re
    url_pattern = r'https?://(?:www\.)?([a-zA-Z0-9.-]+)'
    urls = re.findall(url_pattern, content)
    
    for url_domain in urls:
        # 檢查是否匹配任何已知網域
        for domain, country in domain_to_country.items():
            if domain in url_domain:
                return country
    
    # 如果沒找到 URL，使用文件名稱判斷（原有邏輯）
    return extract_country_from_name(fallback_name)


def extract_country_from_name(name: str) -> str:
    """
    從文件名稱中提取國家名稱（備用方法）
    
    Args:
        name: 文件名稱
        
    Returns:
        國家名稱（中文）
    """
    if not name:
        return " "
        
    name_lower = name.lower()
    
    # 國家關鍵字映射
    country_mapping = {
        '越南': ['越南', 'vietnam', 'vn', 'viet'],
        '泰國': ['泰國', 'thailand', 'thai'],
        '印尼': ['印尼', 'indonesia', 'indonesian'],
        '菲律賓': ['菲律賓', 'philippines', 'philippine', 'filipino'],
        '柬埔寨': ['柬埔寨', 'cambodia', 'cambodian'],
        '新加坡': ['新加坡', 'singapore'],
        '馬來西亞': ['馬來西亞', 'malaysia', 'malaysian'],
        '緬甸': ['緬甸', 'myanmar', 'burma'],
        '寮國': ['寮國', 'laos', 'lao']
    }
    
    # 檢查名稱中是否包含國家關鍵字
    for country, keywords in country_mapping.items():
        for keyword in keywords:
            if keyword in name_lower:
                return country
    
    # 如果沒有匹配到，返回未知
    return " "


def translate_title_to_chinese(title: str) -> str:
    """
    將新聞標題翻譯為繁體中文（如果需要）
    
    Args:
        title: 原始標題
        
    Returns:
        繁體中文標題
    """
    if not title:
        return title
    
    # 簡單檢測：如果標題中中文字符超過 30%，認為已經是中文
    chinese_chars = sum(1 for char in title if '\u4e00' <= char <= '\u9fff')
    if chinese_chars / len(title) > 0.3:
        return title
    
    # 使用批次翻譯函數處理單個標題
    results = batch_translate_titles([title])
    return results.get(title, title)


def batch_translate_titles(titles: List[str]) -> Dict[str, str]:
    """
    批次翻譯多個新聞標題為繁體中文（單次 API 呼叫）
    
    Args:
        titles: 原始標題列表
        
    Returns:
        字典 {原始標題: 翻譯後標題}
    """
    if not titles:
        return {}
    
    # 過濾已經是中文的標題
    titles_to_translate = []
    result = {}
    
    for title in titles:
        if not title:
            result[title] = title
            continue
        chinese_chars = sum(1 for char in title if '\u4e00' <= char <= '\u9fff')
        if len(title) > 0 and chinese_chars / len(title) > 0.3:
            result[title] = title  # 已經是中文，不需翻譯
        else:
            titles_to_translate.append(title)
    
    # 如果所有標題都是中文，直接返回
    if not titles_to_translate:
        return result
    
    try:
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            # 沒有 API key，返回原始標題
            for title in titles_to_translate:
                result[title] = title
            return result
        
        client = OpenAI(api_key=api_key)
        
        # 構建批次翻譯提示
        numbered_titles = "\n".join([f"{i+1}. {t}" for i, t in enumerate(titles_to_translate)])
        
        prompt = f"""請將以下新聞標題翻譯成繁體中文。每行一個標題，保持編號對應。只需回答翻譯結果，不要有其他說明。

原標題列表：
{numbered_titles}

繁體中文翻譯（保持編號格式）："""
        
        response = client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_completion_tokens=500
        )
        
        translated_text = response.choices[0].message.content.strip()
        print(f"📝 批次翻譯完成: {len(titles_to_translate)} 個標題")
        
        # 解析翻譯結果
        lines = translated_text.split("\n")
        for i, title in enumerate(titles_to_translate):
            if i < len(lines):
                # 移除編號前綴 (如 "1. ", "2. " 等)
                translated = re.sub(r'^\d+\.\s*', '', lines[i]).strip()
                result[title] = translated if translated else title
            else:
                result[title] = title
        
        return result
        
    except Exception as e:
        import traceback
        print(f"批次標題翻譯失敗: {e}")
        print(f"錯誤詳情: {traceback.format_exc()}")
        # 翻譯失敗時返回原始標題
        for title in titles_to_translate:
            result[title] = title
        return result


def parse_news_from_content(content: str) -> List[Dict[str, str]]:
    """
    從文件內容中解析新聞列表
    支持兩種格式：
    1. 單篇新聞文檔（NEWS 類型）：包含標題、發布時間、內容和來源
    2. 多篇新聞列表（RESEARCH 類型）：Markdown 格式，標題為 ###
    
    Args:
        content: 文件內容（Markdown 格式）
        
    Returns:
        新聞列表，每個元素包含 title, date, summary, link
    """
    news_items = []
    original_titles = []  # 收集所有原始標題
    
    # 檢測是否為單篇新聞文檔（以 # 標題開頭，包含發布時間和來源）
    if content.strip().startswith('# ') and '**發布時間**' in content and '**來源**' in content:
        # 單篇新聞格式
        news_item = {
            'title': '',
            'original_title': '',  # 暫存原始標題
            'date': '',
            'summary': '',
            'link': ''
        }
        
        lines = content.strip().split('\n')
        
        # 提取標題（第一行，去掉 # 符號）
        if lines:
            original_title = lines[0].replace('#', '').strip()
            news_item['original_title'] = original_title
            original_titles.append(original_title)
        
        # 提取發布時間
        date_pattern = r'\*\*發布時間\*\*[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?)'
        date_match = re.search(date_pattern, content)
        if date_match:
            news_item['date'] = date_match.group(1).replace('年', '-').replace('月', '-').replace('日', '')
        
        # 提取來源 URL
        source_pattern = r'\*\*來源\*\*[：:]\s*(https?://[^\s]+)'
        source_match = re.search(source_pattern, content)
        if source_match:
            news_item['link'] = source_match.group(1).strip()
        
        # 提取摘要（移除標題、發布時間和來源後的內容）
        summary_text = content
        summary_text = re.sub(r'^#[^\n]+\n+', '', summary_text)  # 移除標題
        summary_text = re.sub(r'\*\*發布時間\*\*[：:][^\n]+\n*', '', summary_text)  # 移除發布時間
        summary_text = re.sub(r'發布時間[：:][^\n]+\n*', '', summary_text)  # 移除沒有粗體的發布時間
        summary_text = re.sub(r'\*\*來源\*\*[：:][^\n]+', '', summary_text)  # 移除來源
        # 移除所有 URL
        summary_text = re.sub(r'`https?://[^`]+`', '', summary_text)  # 移除反引號中的 URL
        summary_text = re.sub(r'\[[^\]]*\]\([^\)]*https?://[^\)]*\)', '', summary_text)  # 移除 Markdown 連結
        summary_text = re.sub(r'https?://[^\s\)\]]+', '', summary_text)  # 移除所有其他 URL
        # 移除特殊標記符號和格式
        summary_text = re.sub(r'\*\*[^*]+\*\*[：:]?', '', summary_text)  # 移除粗體標記
        summary_text = re.sub(r'[\[\]\(\)「」『』【】]', '', summary_text)  # 移除各種括號
        summary_text = re.sub(r'[•·▪▸►▶]', '', summary_text)  # 移除列表符號
        # 過濾非中英文字符（保留中文、英文、數字、常用標點）
        summary_text = re.sub(r'[^\u4e00-\u9fff\u3000-\u303fa-zA-Z0-9\s，。！？、；：,.!?\'\"%-]', '', summary_text)
        summary_text = re.sub(r'\n+', ' ', summary_text)  # 合併換行
        summary_text = re.sub(r'\s+', ' ', summary_text)  # 合併空白
        summary_text = summary_text.strip()
        
        news_item['summary'] = summary_text[:500] if summary_text else ''
        
        if news_item['original_title'] and news_item['summary']:
            news_items.append(news_item)
    else:
        # 原有的多篇新聞列表解析邏輯（RESEARCH 類型）
        # 先移除文末的總結區塊
        summary_section_pattern = r'\n##\s+(摘要|期間重點|覆蓋度|缺口|後續建議|Credit Memo).*$'
        content = re.sub(summary_section_pattern, '', content, flags=re.DOTALL)
        
        # 按 ### 標題分割新聞項目
        sections = re.split(r'\n###\s+', content)
        
        for section in sections:
            section = section.strip()
            if not section or len(section) < 20:
                continue
            
            # 跳過非新聞標題
            first_line = section.split('\n')[0].strip()
            if any(keyword in first_line for keyword in ['回覆重點', '越南', '泰國', '印尼', '菲律賓', '柬埔寨', 'Vietnam', 'Thailand', 'Indonesia', 'Philippines', 'Cambodia']):
                continue
            if first_line.startswith('#') or first_line.startswith('【'):
                continue
            
            news_item = {
                'title': '',
                'original_title': '',  # 暫存原始標題
                'date': '',
                'summary': '',
                'link': ''
            }
            
            # 提取標題（第一行）
            lines = section.split('\n')
            if lines:
                original_title = lines[0].strip()
                news_item['original_title'] = original_title
                original_titles.append(original_title)
            
            # 提取發布時間
            date_pattern = r'發布時間[：:]\s*(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?)'
            date_match = re.search(date_pattern, section)
            if date_match:
                news_item['date'] = date_match.group(1).replace('年', '-').replace('月', '-').replace('日', '')
            
            # 提取連結
            link_pattern = r'`(https?://[^`]+)`'
            link_match = re.search(link_pattern, section)
            if link_match:
                news_item['link'] = link_match.group(1).strip()
            else:
                plain_link_match = re.search(r'(https?://[^\s\)]+)', section)
                if plain_link_match:
                    news_item['link'] = plain_link_match.group(1).strip()
            
            # 提取摘要
            summary_text = section
            if lines:
                summary_text = '\n'.join(lines[1:])
            # 移除發布時間（多種格式）
            summary_text = re.sub(r'\*\*發布時間\*\*[：:][^\n]+\n*', '', summary_text)  # 粗體格式
            summary_text = re.sub(r'發布時間[：:][^\n]+\n*', '', summary_text)  # 普通格式
            summary_text = re.sub(r'\d{4}[-/年]\d{1,2}[-/月]\d{1,2}[日]?', '', summary_text)  # 移除日期格式
            summary_text = re.sub(r'---+.*$', '', summary_text, flags=re.DOTALL)
            # 移除所有類型的 URL
            summary_text = re.sub(r'\[([^\]]*)\]\([^\)]*\)', r'\1', summary_text)  # Markdown 連結轉文字
            summary_text = re.sub(r'\([^\)]*https?://[^\)]*\)', '', summary_text)  # 括號內的 URL
            summary_text = re.sub(r'\[[^\]]*\]', '', summary_text)  # 移除剩餘的中括號
            summary_text = re.sub(r'`https?://[^`]+`', '', summary_text)  # 反引號中的 URL
            summary_text = re.sub(r'https?://[^\s\)\]]+', '', summary_text)  # 所有其他 URL
            summary_text = re.sub(r'#+\s*', '', summary_text)
            summary_text = re.sub(r'\(\s*\)', '', summary_text)  # 移除空括號
            # 移除特殊標記符號
            summary_text = re.sub(r'\*\*[^*]+\*\*[：:]?', '', summary_text)  # 移除粗體標記
            summary_text = re.sub(r'[•·▪▸►▶]', '', summary_text)  # 移除列表符號
            # 過濾非中英文字符（保留中文、英文、數字、常用標點）
            summary_text = re.sub(r'[^\u4e00-\u9fff\u3000-\u303fa-zA-Z0-9\s，。！？、；：,.!?\'\"%-]', '', summary_text)
            summary_text = re.sub(r'\n+', ' ', summary_text)
            summary_text = re.sub(r'\s+', ' ', summary_text)
            summary_text = summary_text.strip()
            
            news_item['summary'] = summary_text[:500] if summary_text else ''
            
            # 只有標題和摘要都存在時才加入列表
            if news_item['original_title'] and news_item['summary']:
                news_items.append(news_item)
    
    # 批次翻譯所有標題（單次 API 呼叫）
    if original_titles:
        title_translations = batch_translate_titles(original_titles)
        for item in news_items:
            original = item.get('original_title', '')
            item['title'] = title_translations.get(original, original)
            del item['original_title']  # 移除暫存欄位
    
    return news_items


def generate_news_excel(
    document_name: str,
    document_content: str,
    output_dir: str = "exports"
) -> Dict[str, Any]:
    """
    從文件內容生成新聞 Excel 報告
    
    Args:
        document_name: 文件名稱
        document_content: 文件內容（包含新聞列表）
        output_dir: 輸出目錄
        
    Returns:
        包含 success, filepath, filename, count, news_items 的字典
    """
    try:
        # 確保輸出目錄存在
        os.makedirs(output_dir, exist_ok=True)
        
        # 解析新聞列表
        news_items = parse_news_from_content(document_content)
        
        if not news_items:
            return {
                "success": False,
                "error": "未能從文件中解析出新聞項目"
            }
        
        # 從數據庫記錄獲取國家（不再重複調用 LLM）
        from news_store import news_store
        country = ""  # 默認值改為空字符串
        
        # 嘗試通過文件名查詢數據庫獲取國家
        try:
            all_records = news_store.get_all_records()
            for record in all_records:
                if record.get('name') == document_name:
                    country = record.get('country', '')
                    # 如果是'未知'也視為空
                    if country == ' ':
                        country = ''
                    break
        except Exception as e:
            print(f"⚠️ 從數據庫獲取國家失敗: {e}")
        
        # 如果數據庫中沒有找到，嘗試從 tags 中獲取
        if not country:
            try:
                for record in all_records:
                    if record.get('name') == document_name:
                        tags = record.get('tags', [])
                        if tags and len(tags) > 0:
                            country = tags[0] if tags[0] != ' ' else ''
                        break
            except Exception as e:
                print(f"⚠️ 從 tags 獲取國家失敗: {e}")
        
        for item in news_items:
            item['source_doc'] = country
        
        # 創建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "新聞報告"
        
        # 設定標題行
        headers = ["編號", "新聞標題", "發布時間", "新聞摘要", "新聞連結", "來源國家"]
        ws.append(headers)
        
        # 設定標題樣式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 填入新聞資料
        for idx, news in enumerate(news_items, start=1):
            ws.append([
                idx,
                news.get('title', ''),
                news.get('date', ''),
                news.get('summary', ''),
                news.get('link', ''),
                news.get('source_doc', '')
            ])
        
        # 設定欄寬
        column_widths = {
            'A': 8,   # 編號
            'B': 40,  # 標題
            'C': 15,  # 時間
            'D': 60,  # 摘要
            'E': 50,  # 連結
            'F': 20   # 來源國家
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 設定資料行樣式
        data_alignment = Alignment(vertical="top", wrap_text=True)
        for row in range(2, len(news_items) + 2):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = data_alignment
        
        # 凍結首行
        ws.freeze_panes = "A2"
        
        # 生成檔案名稱
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_doc_name = re.sub(r'[^\w\s-]', '', document_name)[:30]
        filename = f"新聞報告_{safe_doc_name}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # 儲存檔案
        wb.save(filepath)
        
        return {
            "success": True,
            "filepath": filepath,
            "filename": filename,
            "count": len(news_items),
            "news_items": news_items
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": f"生成 Excel 時發生錯誤: {str(e)}"
        }


def cleanup_old_exports(output_dir: str = "exports", max_age_days: int = 7):
    """
    清理超過指定天數的匯出檔案
    
    Args:
        output_dir: 輸出目錄
        max_age_days: 保留天數
    """
    try:
        if not os.path.exists(output_dir):
            return
        
        current_time = datetime.now()
        max_age_seconds = max_age_days * 24 * 60 * 60
        
        for filename in os.listdir(output_dir):
            filepath = os.path.join(output_dir, filename)
            
            if not os.path.isfile(filepath):
                continue
            
            # 檢查檔案年齡
            file_age = current_time.timestamp() - os.path.getmtime(filepath)
            
            if file_age > max_age_seconds:
                try:
                    os.remove(filepath)
                    print(f"已刪除舊檔案: {filename}")
                except Exception as e:
                    print(f"刪除檔案失敗 {filename}: {e}")
    
    except Exception as e:
        print(f"清理舊檔案時發生錯誤: {e}")


def generate_batch_news_excel(
    documents: List[Dict[str, str]],
    output_dir: str = "exports"
) -> Dict[str, Any]:
    """
    批次匯出多個文件的新聞到一個 Excel 檔案
    
    Args:
        documents: 文件列表，每個文件包含 name 和 content
        output_dir: 輸出目錄
        
    Returns:
        包含 success, filepath, filename, count, news_items 的字典
    """
    try:
        os.makedirs(output_dir, exist_ok=True)
        
        # 收集所有新聞項目
        all_news_items = []
        
        for doc in documents:
            doc_name = doc.get('name', '未命名')
            doc_content = doc.get('content', '')
            
            if not doc_content:
                continue
            
            # 解析該文件的新聞
            news_items = parse_news_from_content(doc_content)
            
            # 使用 LLM 判斷國家
            country = extract_country_from_content(doc_content, fallback_name=doc_name)
            for item in news_items:
                item['source_doc'] = country
            
            all_news_items.extend(news_items)
        
        if not all_news_items:
            return {
                "success": False,
                "error": "沒有找到可匯出的新聞項目"
            }
        
        # 創建 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "新聞報告"
        
        # 設定標題行
        headers = ["編號", "新聞標題", "發布時間", "新聞摘要", "新聞連結", "來源國家"]
        ws.append(headers)
        
        # 設定標題樣式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 填入新聞資料
        for idx, news in enumerate(all_news_items, start=1):
            ws.append([
                idx,
                news.get('title', ''),
                news.get('date', ''),
                news.get('summary', ''),
                news.get('link', ''),
                news.get('source_doc', '')
            ])
        
        # 設定欄寬
        column_widths = {
            'A': 8,   # 編號
            'B': 40,  # 標題
            'C': 15,  # 時間
            'D': 60,  # 摘要
            'E': 50,  # 連結
            'F': 20   # 來源國家
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 設定資料行樣式
        data_alignment = Alignment(vertical="top", wrap_text=True)
        for row in range(2, len(all_news_items) + 2):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = data_alignment
        
        # 凍結首行
        ws.freeze_panes = "A2"
        
        # 生成檔案名稱
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"新聞報告_批次匯出_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # 儲存檔案
        wb.save(filepath)
        
        return {
            "success": True,
            "filepath": filepath,
            "filename": filename,
            "count": len(all_news_items),
            "news_items": all_news_items
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": f"批次生成 Excel 時發生錯誤: {str(e)}"
        }
